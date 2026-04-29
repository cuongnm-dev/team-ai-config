#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fill_xlsx_engine.py — Template-schema-driven Excel fill engine.

Reads:
  - template (.xlsx)        — ETC-provided, immutable
  - schema  (.yaml)         — one-time analysis of template structure
  - content-data (.json)    — data produced by tdoc-data-writer

Writes:
  - output (.xlsx) — template copy with cells filled per schema

Guarantees:
  - NEVER writes to a formula cell declared in schema.preserve.formula_cells
  - NEVER writes to a MergedCell that is not the top-left anchor
  - Clears data_table.clear_columns before writing new rows
  - Runs validators after write; aborts (non-zero exit) if any fail

Usage:
  python fill_xlsx_engine.py \
    --template   test-case.xlsx \
    --schema     schemas/test-case.xlsx.schema.yaml \
    --data       content-data.json \
    --output     out/kich-ban-kiem-thu.xlsx \
    [--report    out/xlsx-fill-report.json]
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    import yaml
except ImportError:
    print("ERROR: PyYAML required. Run: pip install pyyaml", file=sys.stderr)
    sys.exit(2)

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# ─────────────────────────── Fill report ───────────────────────────

@dataclass
class FillReport:
    writes_done: int = 0
    writes_skipped_formula: int = 0
    writes_skipped_protected: int = 0
    rows_cleared: int = 0
    rows_written: int = 0
    warnings: list[str] = field(default_factory=list)
    validator_failures: list[str] = field(default_factory=list)
    per_sheet: dict[str, dict[str, int]] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            "writes_done": self.writes_done,
            "writes_skipped_formula": self.writes_skipped_formula,
            "writes_skipped_protected": self.writes_skipped_protected,
            "rows_cleared": self.rows_cleared,
            "rows_written": self.rows_written,
            "warnings": self.warnings,
            "validator_failures": self.validator_failures,
            "per_sheet": self.per_sheet,
            "status": "ok" if not self.validator_failures else "validation_failed",
        }


# ─────────────────────────── JSONPath lite ───────────────────────────

_DOT_RE = re.compile(r"\.(?=[^.\[]|$)")


def resolve(data: dict, path: str | None) -> Any:
    """Resolve a '$.a.b.c' path in `data`. Returns None if any segment missing.

    Supports dots only (no wildcards, no [index] — keep it simple & explicit).
    """
    if path is None:
        return None
    if not path.startswith("$"):
        return path  # literal
    segments = [s for s in path.lstrip("$").lstrip(".").split(".") if s]
    cur: Any = data
    for seg in segments:
        if isinstance(cur, dict) and seg in cur:
            cur = cur[seg]
        else:
            return None
    return cur


# ─────────────────────────── Transforms ───────────────────────────

def apply_transform(value: Any, transform: str, schema: dict) -> Any:
    """Apply a named transform. Returns transformed value or original if unknown."""
    if value is None:
        return None

    if transform == "upper":
        return str(value).upper()
    if transform == "lower":
        return str(value).lower()

    m = re.match(r"truncate\((\d+)\)", transform)
    if m:
        n = int(m.group(1))
        s = str(value)
        return s if len(s) <= n else s[: n - 3] + "..."

    m = re.match(r"join\(['\"](.+?)['\"]\)", transform)
    if m:
        sep = m.group(1)
        if isinstance(value, list):
            return sep.join(str(x) for x in value)
        return str(value)

    if transform == "numbered_join":
        # value expected: list[str] or list[dict{no, text}]
        if isinstance(value, list):
            lines = []
            for i, item in enumerate(value, 1):
                if isinstance(item, dict):
                    no = item.get("no", i)
                    txt = item.get("text") or item.get("action") or item.get("expected") or ""
                    lines.append(f"Bước {no}: {txt}")
                else:
                    lines.append(f"Bước {i}: {item}")
            return "\n".join(lines)
        return str(value)

    if transform == "priority_map":
        pm = schema.get("priority_map", {})
        default = schema.get("priority_default", "Normal")
        return pm.get(str(value), default)

    return value  # unknown transform — pass through


def render_template_string(tpl: str, data: dict) -> str:
    """Replace {a.b.c} tokens with resolve(data, '$.a.b.c')."""
    def sub(m: re.Match) -> str:
        path = "$." + m.group(1)
        v = resolve(data, path)
        return str(v) if v is not None else ""
    return re.sub(r"\{([a-zA-Z0-9_.]+)\}", sub, tpl)


def resolve_write_value(write: dict, data: dict, schema: dict) -> Any:
    """Resolve a write entry's final value.

    Priority:
      1. Hardcoded value       (write.value)
      2. Source from data      (write.source) — PRIMARY for data-driven fields
      3. Template fallback     (write.template) — used only when source missing
      4. Default               (write.default)
    """
    if "value" in write:
        return write["value"]

    # Try source first (data-driven)
    if "source" in write:
        v = resolve(data, write["source"])
        if v is not None and v != "":
            if "transform" in write:
                v = apply_transform(v, write["transform"], schema)
            return v
        # source empty → fall through to template/default

    # Try template (built from data)
    if "template" in write:
        v = render_template_string(write["template"], data)
        if v:
            return v

    return write.get("default")


# ─────────────────────────── Safe cell write ───────────────────────────

def get_merge_anchor(ws, cell_ref: str) -> str:
    """If cell_ref is inside a merged range, return the top-left anchor; else return cell_ref."""
    if not isinstance(ws[cell_ref], MergedCell):
        return cell_ref
    for rng in ws.merged_cells.ranges:
        if cell_ref in rng:
            return f"{get_column_letter(rng.min_col)}{rng.min_row}"
    return cell_ref  # shouldn't happen


def is_formula_cell(cell_ref: str, sheet_schema: dict) -> bool:
    preserve = sheet_schema.get("preserve", {}) or {}
    formula_cells = set(preserve.get("formula_cells") or [])
    return cell_ref in formula_cells


def safe_write(ws, cell_ref: str, value: Any, sheet_schema: dict, report: FillReport) -> bool:
    """Write value to cell unless it's a formula cell. Routes through merge anchor."""
    if is_formula_cell(cell_ref, sheet_schema):
        report.writes_skipped_formula += 1
        report.warnings.append(
            f"Skipped formula-cell write: {ws.title}!{cell_ref} (would destroy formula)"
        )
        return False
    anchor = get_merge_anchor(ws, cell_ref)
    if is_formula_cell(anchor, sheet_schema):
        report.writes_skipped_formula += 1
        return False
    ws[anchor] = value
    report.writes_done += 1
    return True


# ─────────────────────────── Sheet fill ───────────────────────────

def process_sheet_writes(ws, sheet_schema: dict, data: dict, schema: dict, report: FillReport):
    """Process sheet.writes + sheet.clear_cells."""
    sheet_stats = report.per_sheet.setdefault(ws.title, {"writes": 0, "rows": 0})

    # Clear first
    for cell_ref in sheet_schema.get("clear_cells", []) or []:
        if is_formula_cell(cell_ref, sheet_schema):
            continue
        anchor = get_merge_anchor(ws, cell_ref)
        ws[anchor] = None

    # Writes
    for write in sheet_schema.get("writes", []) or []:
        cell_ref = write["cell"]
        value = resolve_write_value(write, data, schema)
        if value is None:
            continue
        if safe_write(ws, cell_ref, value, sheet_schema, report):
            sheet_stats["writes"] += 1


def _unmerge_data_region(ws, start_row: int, end_row: int, report: FillReport) -> int:
    """Unmerge any merged range that falls within [start_row, end_row].

    Template often has sample data with feature-group header merges (A:M).
    Those merges must be removed before we fill flat per-TC data;
    otherwise writing to cells inside the merge is silently skipped.

    Returns count of unmerged ranges.
    """
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        # overlap check
        if rng.min_row >= start_row and rng.max_row <= end_row:
            to_unmerge.append(str(rng))
    for r in to_unmerge:
        ws.unmerge_cells(r)
    if to_unmerge:
        report.warnings.append(
            f"{ws.title}: unmerged {len(to_unmerge)} ranges in data region "
            f"rows {start_row}-{end_row} (template sample-data leftovers)"
        )
    return len(to_unmerge)


def process_data_table(ws, sheet_schema: dict, data: dict, schema: dict, report: FillReport):
    """Clear data columns then write rows from data_table.source.

    Steps:
      1. Unmerge any merged range entirely within data region (template
         sample-data leftovers). Without this, writes into merged cells
         are silently skipped → data loss.
      2. Clear clear_columns rows [start_row .. end_cap] (formula-safe).
      3. Write rows from source into consecutive rows starting at start_row.
    """
    dt = sheet_schema.get("data_table")
    if not dt:
        return
    start_row = dt["start_row"]
    end_cap = dt.get("end_row_hard_cap", 600)
    clear_cols = dt.get("clear_columns", [])

    rows_data = resolve(data, dt["source"]) or []
    if not isinstance(rows_data, list):
        report.warnings.append(
            f"{ws.title}: data_table.source '{dt['source']}' resolved to non-list; skipped"
        )
        return

    # ── STEP 1: Unmerge template sample-data merges in data region ──
    _unmerge_data_region(ws, start_row, end_cap, report)

    # ── STEP 2: Clear clear_columns in data region ──
    protected = set(sheet_schema.get("preserve", {}).get("protected_columns_in_data") or [])
    safe_clear_cols = [c for c in clear_cols if c not in protected]

    for row_num in range(start_row, end_cap + 1):
        for col_letter in safe_clear_cols:
            cell_ref = f"{col_letter}{row_num}"
            if is_formula_cell(cell_ref, sheet_schema):
                continue
            # After unmerge, MergedCell shouldn't exist in this region;
            # keep the check defensively for cross-region merges.
            if isinstance(ws[cell_ref], MergedCell):
                continue
            ws[cell_ref] = None
    report.rows_cleared += (end_cap - start_row + 1)

    # ── STEP 3: Write rows ──
    sheet_stats = report.per_sheet.setdefault(ws.title, {"writes": 0, "rows": 0})
    for i, row_item in enumerate(rows_data):
        target_row = start_row + i
        if target_row > end_cap:
            report.warnings.append(
                f"{ws.title}: row {target_row} exceeds end_row_hard_cap={end_cap}; "
                f"{len(rows_data) - i} rows skipped"
            )
            break
        for col_letter, col_spec in dt["row_template"].items():
            cell_ref = f"{col_letter}{target_row}"
            if col_letter in protected:
                continue
            value = resolve_write_value(col_spec, row_item, schema)
            if value is None:
                continue
            if isinstance(ws[cell_ref], MergedCell):
                report.warnings.append(
                    f"{ws.title}!{cell_ref}: still MergedCell after unmerge; skipped"
                )
                continue
            safe_write(ws, cell_ref, value, sheet_schema, report)
        sheet_stats["rows"] += 1
        report.rows_written += 1


# ─────────────────────────── Validators ───────────────────────────

def run_validators(wb, schema: dict, report: FillReport):
    for v in schema.get("validators", []) or []:
        vtype = v["type"]
        name = v.get("name", vtype)
        try:
            if vtype == "sheet_names_match":
                expected = set(v["expected"])
                actual = set(wb.sheetnames)
                missing = expected - actual
                if missing:
                    report.validator_failures.append(f"{name}: missing sheets {missing}")

            elif vtype == "formula_cells_intact":
                ws = wb[v["sheet"]]
                for cell_ref in v["cells"]:
                    val = ws[cell_ref].value
                    if not (isinstance(val, str) and val.startswith("=")):
                        report.validator_failures.append(
                            f"{name}: {v['sheet']}!{cell_ref} is not a formula (got: {val!r})"
                        )

            elif vtype == "cell_values_in_set":
                ws = wb[v["sheet"]]
                col = v["column"]
                allowed = set(v["allowed"])
                for row in range(v["start_row"], v["end_row"] + 1):
                    val = ws[f"{col}{row}"].value
                    if val not in allowed:
                        report.validator_failures.append(
                            f"{name}: {v['sheet']}!{col}{row} = {val!r} not in allowed set"
                        )
                        break  # only report first occurrence

            elif vtype == "cells_not_equal":
                ws = wb[v["sheet"]]
                for chk in v["checks"]:
                    val = ws[chk["cell"]].value
                    if val == chk["must_not_equal"]:
                        report.validator_failures.append(
                            f"{name}: {v['sheet']}!{chk['cell']} still equals template placeholder"
                        )

            elif vtype == "row_count_min":
                ws = wb[v["sheet"]]
                count = 0
                for row in ws.iter_rows(
                    min_row=v["start_row"],
                    max_row=v["start_row"] + 1000,
                    min_col=column_index_from_string(v["column"]),
                    max_col=column_index_from_string(v["column"]),
                ):
                    if row[0].value:
                        count += 1
                if count < v["min_count"]:
                    # Soft: if no rows at all, it's a warning (possibly no test_cases
                    # data provided). Hard fail only when partial rows present but below min.
                    msg = (f"{name}: {v['sheet']} col {v['column']} has {count} filled rows, "
                           f"expected >= {v['min_count']}")
                    if count == 0 and v.get("soft_if_empty", True):
                        report.warnings.append(msg + " (no data provided — soft warning)")
                    else:
                        report.validator_failures.append(msg)
            else:
                report.warnings.append(f"Unknown validator type: {vtype}")
        except Exception as e:
            report.validator_failures.append(f"{name}: validator crashed: {e}")


# ─────────────────────────── Main ───────────────────────────

def fill(template_path: Path, schema_path: Path, data_path: Path, output_path: Path) -> FillReport:
    report = FillReport()

    # Load
    schema = yaml.safe_load(schema_path.read_text(encoding="utf-8"))
    data = json.loads(data_path.read_text(encoding="utf-8"))

    # Copy template first so we never mutate original
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path, data_only=False, keep_vba=False, keep_links=True)

    # Pre-flight: required sheets
    required = schema.get("required_sheets", [])
    missing = [s for s in required if s not in wb.sheetnames]
    if missing:
        report.validator_failures.append(f"Missing required sheets: {missing}")
        return report

    # Process each declared sheet
    for sheet_name, sheet_schema in (schema.get("sheets") or {}).items():
        if sheet_name not in wb.sheetnames:
            report.warnings.append(f"Schema declares sheet '{sheet_name}' but not in workbook")
            continue
        ws = wb[sheet_name]
        process_sheet_writes(ws, sheet_schema, data, schema, report)
        if "data_table" in sheet_schema:
            process_data_table(ws, sheet_schema, data, schema, report)

    wb.save(output_path)

    # Validate saved workbook (reload to verify on-disk state)
    wb_check = load_workbook(output_path, data_only=False, keep_links=True)
    run_validators(wb_check, schema, report)

    return report


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--schema", required=True)
    ap.add_argument("--data", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--report", default=None)
    args = ap.parse_args()

    report = fill(
        Path(args.template),
        Path(args.schema),
        Path(args.data),
        Path(args.output),
    )

    report_dict = report.to_dict()
    if args.report:
        Path(args.report).parent.mkdir(parents=True, exist_ok=True)
        Path(args.report).write_text(
            json.dumps(report_dict, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    # Human-readable summary
    print(f"Fill: {report.writes_done} cells, {report.rows_written} data rows")
    if report.warnings:
        print(f"Warnings: {len(report.warnings)}")
        for w in report.warnings[:5]:
            print(f"  - {w}")
    if report.validator_failures:
        print(f"VALIDATION FAILURES: {len(report.validator_failures)}")
        for f in report.validator_failures:
            print(f"  ✗ {f}")
        sys.exit(1)
    print("OK")


if __name__ == "__main__":
    main()
