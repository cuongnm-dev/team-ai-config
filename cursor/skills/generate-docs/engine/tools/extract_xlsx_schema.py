#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract_xlsx_schema.py — Analyze an Excel template and emit a structural report.

Output: JSON describing every sheet's:
  - dimensions
  - merged cells
  - formula cells (address + formula string)
  - data validations (ranges + formula1)
  - conditional formatting rules
  - non-empty static cells (first 30 rows for inspection)
  - defined styles / number formats

Purpose: feed this report to a human reviewer (or LLM one time) to produce
test-case-schema.yaml for fill_xlsx_engine.py.

Usage:
  python extract_xlsx_schema.py <template.xlsx> [--out schema-report.json]
"""
from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


def extract_sheet(ws) -> dict:
    info = {
        "name": ws.title,
        "dimensions": ws.dimensions,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
        "formulas": [],
        "data_validations": [],
        "conditional_formatting": [],
        "non_empty_preview": [],
        "column_widths": {},
        "row_heights": {},
    }

    # Column widths
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width is not None:
            info["column_widths"][col_letter] = round(dim.width, 2)

    # Row heights
    for row_num, dim in ws.row_dimensions.items():
        if dim.height is not None:
            info["row_heights"][str(row_num)] = round(dim.height, 2)

    # Walk cells (bounded to max_row/col to avoid scanning blank region)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 600)):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            # Formulas
            if isinstance(val, str) and val.startswith("="):
                info["formulas"].append({
                    "cell": cell.coordinate,
                    "formula": val,
                })
            elif val is not None and cell.row <= 30:
                # Preview static content (first 30 rows)
                info["non_empty_preview"].append({
                    "cell": cell.coordinate,
                    "value": str(val)[:200],
                    "style": cell.style,
                    "number_format": cell.number_format,
                })

    # Data validations
    for dv in ws.data_validations.dataValidation:
        info["data_validations"].append({
            "type": dv.type,
            "formula1": dv.formula1,
            "formula2": dv.formula2,
            "ranges": [str(r) for r in dv.sqref.ranges],
            "allow_blank": dv.allowBlank,
            "error_title": dv.errorTitle,
        })

    # Conditional formatting
    for cf_range, rules in ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            info["conditional_formatting"].append({
                "range": str(cf_range),
                "type": rule.type,
                "operator": getattr(rule, "operator", None),
                "formula": list(rule.formula) if rule.formula else [],
                "priority": rule.priority,
            })

    return info


def extract_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=False, keep_vba=False, keep_links=True)
    report = {
        "template_path": str(path),
        "sheet_names": wb.sheetnames,
        "defined_names": {},
        "sheets": [],
    }
    for dn in wb.defined_names:
        dnobj = wb.defined_names[dn]
        report["defined_names"][dn] = str(dnobj.value) if hasattr(dnobj, "value") else str(dnobj)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        report["sheets"].append(extract_sheet(ws))

    return report


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("template", help="Path to .xlsx template")
    ap.add_argument("--out", default=None, help="Output JSON path (default: stdout)")
    args = ap.parse_args()

    path = Path(args.template)
    if not path.exists():
        print(f"ERROR: file not found: {path}", file=sys.stderr)
        sys.exit(1)

    report = extract_workbook(path)
    txt = json.dumps(report, ensure_ascii=False, indent=2)
    if args.out:
        Path(args.out).write_text(txt, encoding="utf-8")
        print(f"Wrote {args.out}  ({len(report['sheets'])} sheets, "
              f"{sum(len(s['formulas']) for s in report['sheets'])} formulas)")
    else:
        print(txt)


if __name__ == "__main__":
    main()
