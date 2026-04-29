#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fill-testcase.py — Fill test-case.xlsx template from test-cases.json

Fills ALL 5 sheets:
  - Cover:        revision history row
  - Overview:     project info (name, code, version, client)
  - Test Result:  project metadata (name, code, env, version, tester, date range)
  - Tên chức năng: UI test cases
  - Tên API:      API test cases

Usage:
  python fill-testcase.py \
    --template     /path/docs/templates/test-case.xlsx \
    --json-in      /path/docs/generated/output/test-cases.json \
    --output       /path/docs/generated/output/kich-ban-kiem-thu.xlsx \
    --project-name "Tên Dự Án" \
    --client-name  "Tên Khách Hàng" \
    --dev-unit     "Tên Đơn Vị Phát Triển"

Template constraints (verified against test-case.xlsx v1.0):
  - Sheet "Tên chức năng": KEEP NAME — Test Result R14 references it with formula
  - Sheet "Tên API":        KEEP NAME — Test Result R15 references it with formula
  - Col A in data rows:     auto-formula =IF(OR(B{row}...) — NEVER fill
  - UI test cases start at: row 11
  - API test cases start at: row 14
  - Overview values at col C (merged C5:H5, C6:H6, etc.)
"""

import argparse
import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# 4-level priority mapping (matches Overview sheet definition)
PRIORITY_MAP = {
    "Rất cao":   "Critical",
    "Cao":       "Major",
    "Trung bình":"Normal",
    "Thấp":      "Minor",
}


# ── Sheet fillers ─────────────────────────────────────────────────────────────

def fill_overview(wb, project_name, client_name, dev_unit, today):
    """Fill Overview sheet — project metadata + revision history first row."""
    if "Overview" not in wb.sheetnames:
        return
    ws = wb["Overview"]
    # Merged cells A:B = label, C:H = value
    ws["C5"] = project_name          # Tên dự án
    ws["C6"] = ""                    # Mã dự án — human fills
    ws["C7"] = "1.0"                 # Phiên bản
    ws["C8"] = client_name           # Khách hàng (FIXED: was dev_unit)
    ws["C9"] = f"Hệ thống {project_name}"  # Khái quát về dự án
    ws["C10"] = "[CẦN BỔ SUNG: stack list]"  # Công nghệ sử dụng

    # Revision history row 15
    ws["A15"] = today
    ws["B15"] = "1.0"
    ws["C15"] = "Tạo mới bộ test case"
    ws["E15"] = "Toàn bộ tài liệu"
    ws["G15"] = dev_unit


def fill_cover_history(wb, today, dev_unit):
    if "Cover" not in wb.sheetnames:
        return
    ws = wb["Cover"]
    # Clear placeholder row 16 (<dd/mm/yyyy>, <x.y>, etc.)
    from openpyxl.cell.cell import MergedCell
    for c in range(1, 9):
        cell = ws.cell(row=16, column=c)
        if not isinstance(cell, MergedCell):
            cell.value = None

    # Fill first actual entry at row 17
    ws["A17"] = today
    ws["B17"] = "1.0"
    ws["C17"] = "Tạo mới"
    ws["G17"] = dev_unit
    ws["H17"] = "[CẦN BỔ SUNG: Reviewer]"


def fill_test_result_sheet(wb, project_name, total_tc_ui, total_tc_api, today):
    """Fill 'Test Result' sheet — project metadata."""
    if "Test Result" not in wb.sheetnames:
        return
    ws = wb["Test Result"]
    # Project metadata (merged C:)
    ws["C3"] = project_name                             # Tên dự án
    ws["C4"] = ""                                       # Mã dự án — human
    ws["C5"] = "Web + API (localhost Docker)"           # Môi trường kiểm tra
    ws["C6"] = "1.0"                                    # Version tổng hợp kết quả
    ws["C7"] = "Xem tài liệu: bo-test-case.md, thiet-ke-co-so.md"  # Tài liệu SRS/URD
    ws["C8"] = "[CẦN BỔ SUNG: Tên tester]"              # Người thực hiện
    ws["C9"] = f"{today} - {today}"                     # Thời gian — human updates

    # Note: rows 14-15 (B="Tên chức năng", B="Tên hàm") have formulas
    # referencing 'Tên chức năng'!G7 and 'Tên API'!G7 which auto-count.
    # No manual fill needed.


def write_tc_rows(ws, test_cases, start_row):
    """
    Write test case data. NEVER write col A (auto-formula).
    Col mapping:
      B=Mục đích  C=Các bước  D=Kết quả mong muốn
      E=Checklog  F=Chuyển hướng  G=ID Bug (empty)
      H=Mức độ nghiêm trọng
      I=Kết quả lần 1  J=Trạng thái (leave for tester)
      K=Kết quả lần 2  L=Trạng thái (leave for tester)
      M=Ghi chú
    """
    row = start_row
    for tc in test_cases:
        steps = tc.get("steps", [])
        steps_text = "\n".join(f"Bước {s['no']}: {s['action']}" for s in steps)
        expected_text = "\n".join(
            f"Bước {s['no']}: {s['expected']}" for s in steps if s.get("expected")
        )

        ws.cell(row=row, column=2).value = tc.get("name", "")
        ws.cell(row=row, column=3).value = steps_text
        ws.cell(row=row, column=4).value = expected_text
        ws.cell(row=row, column=5).value = tc.get("checklog", "")
        ws.cell(row=row, column=6).value = tc.get("redirect", "")
        # col 7 = ID Bug — leave empty (tester fills during execution)
        ws.cell(row=row, column=8).value = PRIORITY_MAP.get(
            tc.get("priority", ""), "Normal"
        )
        # cols 9-12 = Kết quả lần 1/2 — leave for tester
        ws.cell(row=row, column=13).value = tc.get("notes", "")
        row += 1
    return row


def fill_ui_sheet(wb, features, project_name, today):
    """Fill 'Tên chức năng' with all UI test cases, grouped by feature."""
    if "Tên chức năng" not in wb.sheetnames:
        return 0
    ws = wb["Tên chức năng"]
    ws["B1"] = project_name
    ws["B2"] = "Kiểm tra các chức năng hệ thống"
    ws["B3"] = "[CẦN BỔ SUNG: Tester]"
    ws["B4"] = today

    # Clear template placeholder rows (cols A-M, rows 11 onwards)
    from openpyxl.cell.cell import MergedCell
    for r in range(11, ws.max_row + 1):
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    row = 11
    total_tc = 0
    for feature in features:
        tcs = feature.get("test_cases", [])
        if not tcs:
            continue
        # Feature header row
        ws.cell(row=row, column=1).value = feature.get("name", "")
        row += 1
        row = write_tc_rows(ws, tcs, row)
        total_tc += len(tcs)
    return total_tc


def fill_api_sheet(wb, features, project_name, today):
    """Fill 'Tên API' with all API test cases. Data starts at row 14."""
    if "Tên API" not in wb.sheetnames:
        return 0
    ws = wb["Tên API"]
    ws["B1"] = project_name
    ws["B2"] = "Kiểm tra các API hệ thống"
    ws["B3"] = "[CẦN BỔ SUNG: Tester]"
    ws["B4"] = today

    # Clear template placeholder rows (cols A-M, rows 14 onwards)
    from openpyxl.cell.cell import MergedCell
    for r in range(14, ws.max_row + 1):
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    row = 14
    total_tc = 0
    for feature in features:
        tcs = feature.get("test_cases", [])
        if not tcs:
            continue
        ws.cell(row=row, column=1).value = feature.get("name", "")
        row += 1
        row = write_tc_rows(ws, tcs, row)
        total_tc += len(tcs)
    return total_tc


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Fill test-case Excel template")
    parser.add_argument("--template", required=True)
    parser.add_argument("--json-in", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--project-name", required=True)
    parser.add_argument("--client-name", default="", help="Tên khách hàng")
    parser.add_argument("--dev-unit", required=True, help="Đơn vị phát triển")
    args = parser.parse_args()

    today = datetime.now().strftime("%d/%m/%Y")
    client_name = args.client_name or args.project_name

    for path, label in [(args.template, "template"), (args.json_in, "json-in")]:
        if not os.path.exists(path):
            print(f"ERROR: {label} not found: {path}")
            return 1

    with open(args.json_in, encoding="utf-8") as f:
        data = json.load(f)

    features = data.get("features", [])
    ui_features = [f for f in features if not f.get("is_api")]
    api_features = [f for f in features if f.get("is_api")]

    # Scalability warning
    total_tc = sum(len(f.get("test_cases", [])) for f in features)
    if total_tc > 500:
        print(f"⚠️ LARGE TEST SUITE: {total_tc} test cases")
        print("   Excel template has row limits (UI: ~500 rows, API: ~500 rows)")
        print("   Consider splitting into multiple .xlsx files per module")

    wb = load_workbook(args.template)

    # Validate required sheets (Issue #13 fix)
    REQUIRED_SHEETS = ["Cover", "Overview", "Test Result", "Tên chức năng", "Tên API"]
    missing_sheets = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing_sheets:
        print(f"⚠️ TEMPLATE VERSION MISMATCH")
        print(f"   Missing sheets: {missing_sheets}")
        print(f"   Available sheets: {wb.sheetnames}")
        print(f"   Script will fill what it can but output may be incomplete.")

    filled_sheets = []
    skipped_sheets = []

    # Fill each sheet only if it exists (graceful degradation)
    if "Cover" in wb.sheetnames:
        fill_cover_history(wb, today, args.dev_unit)
        filled_sheets.append("Cover")
    else:
        skipped_sheets.append("Cover")

    if "Overview" in wb.sheetnames:
        fill_overview(wb, args.project_name, client_name, args.dev_unit, today)
        filled_sheets.append("Overview")
    else:
        skipped_sheets.append("Overview")

    total_ui_tc = fill_ui_sheet(wb, ui_features, args.project_name, today) \
        if "Tên chức năng" in wb.sheetnames else 0
    if "Tên chức năng" in wb.sheetnames:
        filled_sheets.append("Tên chức năng")
    else:
        skipped_sheets.append("Tên chức năng")

    total_api_tc = fill_api_sheet(wb, api_features, args.project_name, today) \
        if "Tên API" in wb.sheetnames else 0
    if "Tên API" in wb.sheetnames:
        filled_sheets.append("Tên API")
    else:
        skipped_sheets.append("Tên API")

    if "Test Result" in wb.sheetnames:
        fill_test_result_sheet(wb, args.project_name, total_ui_tc, total_api_tc, today)
        filled_sheets.append("Test Result")
    else:
        skipped_sheets.append("Test Result")

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)

    size_mb = os.path.getsize(args.output) / (1024 * 1024)
    print(f"Saved: {args.output} ({size_mb:.1f} MB)")
    print(f"  UI test cases:  {total_ui_tc}")
    print(f"  API test cases: {total_api_tc}")
    print(f"  Total:          {total_ui_tc + total_api_tc}")
    print(f"  Sheets filled:  {', '.join(filled_sheets)}")
    if skipped_sheets:
        print(f"  Sheets skipped: {', '.join(skipped_sheets)}")

    # Depth enforcement check (Issue #4 lite — on exporter side)
    if features:
        avg_tc = (total_ui_tc + total_api_tc) / max(len(features), 1)
        if avg_tc < 3:
            print(f"\n⚠️ DEPTH WARNING: avg {avg_tc:.1f} TC/feature (target ≥3)")
            print(f"   {len(features)} features but only {total_ui_tc + total_api_tc} TCs total")
            print(f"   tdoc-testcase-writer may have produced sparse output — regenerate recommended")

    # Write fill report for exporter
    report_path = os.path.join(os.path.dirname(args.output), "testcase-fill-report.json")
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump({
            "generated_at": datetime.now().isoformat(),
            "output_file": args.output,
            "size_mb": round(size_mb, 2),
            "ui_tc": total_ui_tc,
            "api_tc": total_api_tc,
            "total_tc": total_ui_tc + total_api_tc,
            "avg_tc_per_feature": round(avg_tc, 2) if features else 0,
            "filled_sheets": filled_sheets,
            "skipped_sheets": skipped_sheets,
            "missing_required_sheets": missing_sheets,
            "depth_warning": avg_tc < 3 if features else False,
        }, f, ensure_ascii=False, indent=2)

    return 0


if __name__ == "__main__":
    sys.exit(main())
